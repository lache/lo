import re
import struct
from openpyxl import load_workbook

def convert_coordinates(val, negative_sign):
	degrees = int(val[:-2])
	minutes = int(val[-2:])
	v = degrees + minutes / 60
	if negative_sign:
		v = -v
	return v

def export_workbook(sh, fout):
	row = 1
	written = 0
	while True:
		row = row + 1
		locode = sh.cell(row=row,column=2).value
		if locode == None:
			break
		locode = "".join(locode.split())
		coordinates = sh.cell(row=row,column=10).value
		if coordinates == None:
			continue
		coordinates = coordinates_pattern.findall(coordinates)[0]
		lat = convert_coordinates(coordinates[0], coordinates[1] == 'S')
		lng = convert_coordinates(coordinates[2], coordinates[3] == 'W')
		function = sh.cell(row=row,column=6).value
		if not function.startswith('1'):
			continue
		name = "".join(sh.cell(row=row,column=3).value.split())
		locode_bytes = bytes(locode, "utf-8")
		name_bytes = bytes(name, "utf-8")
		#print(len(name_bytes))
		fout.write(struct.pack('<8s64sff', locode_bytes, name_bytes, lat, lng))
		print(locode, coordinates, name, lat, lng)
		written = written + 1
	return written

coordinates_pattern = re.compile('\s*(\d+)([NS])\s+(\d+)([EW])\s*')
wb = load_workbook('./seaports.xlsx')

total_written = 0
with open("seaports.dat", "wb") as fout:
	written_kr = export_workbook(wb['KR'], fout)
	written_jp = export_workbook(wb['JP'], fout)
	written_cn = export_workbook(wb['CN'], fout)
	total_written = written_kr + written_jp + written_cn
	
print('%d entries written.' % total_written)